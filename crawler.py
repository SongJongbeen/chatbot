from selenium import webdriver
import openpyxl

driver=webdriver.Chrome("chromedriver.exe")
driver.implicitly_wait(1)

#go to the login page & login

driver.get("https://everytime.kr/login")
driver.find_element_by_name("userid").send_keys("thdwhdqls")
driver.find_element_by_name("password").send_keys("1041489LM")
driver.find_element_by_xpath('//*[@class="submit"]/input').click()
driver.implicitly_wait(1)

results=[]
cnt=0

while True:
    print("Page "+str(cnt))

    if cnt > 162:   #163페이지까지가 2019년도 게시글
        break
    cnt += 1

    driver.get("https://everytime.kr/382452/p/"+str(cnt))
    driver.implicitly_wait(1)

    #get articles link
    posts=driver.find_elements_by_css_selector("article > a.article")
    links=[post.get_attribute("href") for post in posts]

    #get detail article
    for link in links:
        driver.get(link)

        #댓글
        comments=driver.find_elements_by_css_selector("div.comments > article.parent > p.large") #div.comments만 하면 익명,대댓글,공강,쪽지,댓글단 시간 다 나옴;;;

        for comment in comments:
            results.append(comment.text)

wb = openpyxl.Workbook()
sheet1 = wb.create_sheet('eta_assistant', 1)

for i in range(len(results)):
    sheet1.cell(row=i+1, column=1).value = results[i]

wb.save('eta_assistant_results.xlsx')
