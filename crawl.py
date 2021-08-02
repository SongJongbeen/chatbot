from selenium import webdriver
import openpyxl

driver = webdriver.Chrome("chromedriver.exe")
driver.implicitly_wait(1)

# go to the login page & login

driver.get("https://everytime.kr/login")
driver.find_element_by_name("userid").send_keys("thdwhdqls")
driver.find_element_by_name("password").send_keys("1041489LM")
driver.find_element_by_xpath('//*[@class="submit"]/input').click()
driver.implicitly_wait(1)

main_results = []  # 본문
comment_results = []  # 댓글
main_urls = []
comment_urls = []
cnt = 0

while True:
    print("Page " + str(cnt))

    if cnt > 1:  # 163페이지까지가 2019년도 게시글
        break
    cnt = cnt + 1

    driver.get("https://everytime.kr/382452/p/" + str(cnt))
    driver.implicitly_wait(1)

    # get articles link
    posts = driver.find_elements_by_css_selector("article > a.article")
    links = [post.get_attribute("href") for post in posts]

    # get detail article
    for link in links:
        driver.get(link)

        # 본문
        main_articles = driver.find_elements_by_css_selector("p.large")

        comments = driver.find_elements_by_css_selector("div.comments > article.parent > p.large")  # 댓글(가장 상위에 있는)
        child_comments = driver.find_elements_by_css_selector("div.comments > article.child > p.large")  # 대댓글

        for main_article in main_articles:  # 본문만 추출할 수 있는 html경로를 파악하지 못 해서 본문과 댓글을 우선 전체 크롤링
            main_results.append(main_article.text)  # main_results -> 본문+댓글 리스트
            main_urls.append(link)

        for comment in comments:
            comment_results.append(comment.text)  # comment_results -> 댓글 리스트
            comment_urls.append(link)

        for child_comment in child_comments:  # 대댓글도 comment_results 리스트에 포함시킴
            comment_results.append(child_comment.text)
            comment_urls.append(link)


filter_data = [x for x in main_results if
               all(y not in x for y in comment_results)]  # filter_data에는 '본문'만 담고 있음

# print(comment_results)
# print(filter_data)

wb = openpyxl.Workbook()
sheet1 = wb.create_sheet('eta_assistant', 1)

for i in range(len(main_results)):
    sheet1.cell(row=i+1, column=1).value = main_results[i]
    sheet1.cell(row=i+1, column=2).value = main_urls[i]

wb.save('eta_assistant_results.xlsx')
